"""저장된 엑셀을 다시 열어 DB 조회 결과와 셀 단위로 대조하는 검증기.

xlsx 저장 규칙상 발생하는 정당한 변환만 허용한다:
  - 빈 문자열/None → 빈 셀
  - date → 자정 datetime (엑셀 날짜 직렬화)
  - 제어문자 제거 (엑셀 스펙상 저장 불가 문자)
  - 숫자는 IEEE double 정밀도 (xlsx 한계)
그 외 모든 차이는 불일치로 보고한다.
"""

import datetime
import decimal
import logging
import math

from openpyxl import load_workbook

from excel_writer import EXCEL_MAX_ROWS, ILLEGAL_CHARACTERS_RE

log = logging.getLogger("jdbc2excel")


def verify_workbook(path, sheet_map):
    """write_workbook이 돌려준 (시트명, QueryResult) 목록과 파일 내용을 대조한다.

    불일치 설명 문자열 목록을 돌려준다. 빈 목록이면 검증 통과.
    """
    mismatches = []
    wb = load_workbook(path, read_only=True)
    try:
        for sheet_name, res in sheet_map:
            if res.error is not None:
                continue
            if sheet_name not in wb.sheetnames:
                mismatches.append(f"'{sheet_name}': 시트가 파일에 없음")
                continue
            _verify_sheet(wb[sheet_name], sheet_name, res, mismatches)
    finally:
        wb.close()
    return mismatches


def _verify_sheet(ws, sheet_name, res, mismatches):
    expected_rows = res.rows[: EXCEL_MAX_ROWS - 1]
    row_iter = ws.iter_rows(values_only=True)

    header = next(row_iter, None) or ()
    for c, exp in enumerate(res.columns):
        actual = header[c] if c < len(header) else None
        if not _cells_match(exp, actual):
            mismatches.append(
                f"'{sheet_name}' 헤더 {c + 1}열: 기대 {exp!r} ≠ 실제 {actual!r}"
            )

    n_actual = 0
    for actual_row in row_iter:
        if n_actual < len(expected_rows):
            exp_row = expected_rows[n_actual]
            for c, exp in enumerate(exp_row):
                actual = actual_row[c] if c < len(actual_row) else None
                if not _cells_match(exp, actual):
                    mismatches.append(
                        f"'{sheet_name}' {n_actual + 2}행 {c + 1}열: "
                        f"기대 {exp!r} ≠ 실제 {actual!r}"
                    )
            for c in range(len(exp_row), len(actual_row)):
                if actual_row[c] is not None:
                    mismatches.append(
                        f"'{sheet_name}' {n_actual + 2}행 {c + 1}열: "
                        f"예상 밖의 값 {actual_row[c]!r}"
                    )
        elif any(v is not None for v in actual_row):
            mismatches.append(f"'{sheet_name}' {n_actual + 2}행: 예상 밖의 추가 행")
        n_actual += 1

    if n_actual < len(expected_rows):
        mismatches.append(
            f"'{sheet_name}': 데이터 행 누락 (기대 {len(expected_rows)}행, 실제 {n_actual}행)"
        )


def _cells_match(expected, actual):
    if expected is None:
        return actual is None
    if isinstance(expected, str):
        cleaned = ILLEGAL_CHARACTERS_RE.sub("", expected)
        if cleaned == "":
            return actual is None or actual == ""
        return actual == cleaned
    if isinstance(expected, bool):
        return actual == expected
    if isinstance(expected, (int, float, decimal.Decimal)):
        if isinstance(actual, bool) or not isinstance(actual, (int, float)):
            return False
        return math.isclose(float(expected), float(actual), rel_tol=1e-12, abs_tol=1e-12)
    if isinstance(expected, datetime.datetime):
        return (
            isinstance(actual, datetime.datetime)
            and abs((actual - expected).total_seconds()) < 0.001
        )
    if isinstance(expected, datetime.date):
        return actual == expected or actual == datetime.datetime.combine(
            expected, datetime.time.min
        )
    if isinstance(expected, datetime.time):
        return actual == expected
    return actual is not None and str(expected) == str(actual)

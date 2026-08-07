"""openpyxl 기반 엑셀 생성.

시트 서식: 헤더 볼드+회색 배경+자동필터, 전체 셀 검정 실선 테두리+자동 줄바꿈,
첫 행 고정, 열 너비 자동(한글 등 전각 문자는 2칸으로 계산).
"""

import logging
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except ImportError:
    ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010\013\014\016-\037]")

log = logging.getLogger("jdbc2excel")

EXCEL_MAX_ROWS = 1_048_576
MAX_SHEET_NAME = 31
MIN_COL_WIDTH = 8
MAX_COL_WIDTH = 60
WIDTH_SAMPLE_ROWS = 500

_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")

_THIN_BLACK = Side(style="thin", color="000000")
BORDER = Border(left=_THIN_BLACK, right=_THIN_BLACK, top=_THIN_BLACK, bottom=_THIN_BLACK)
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")
CELL_ALIGN = Alignment(wrap_text=True, vertical="top")
ERROR_FONT = Font(bold=True, color="CC0000")


def write_workbook(path, results):
    """QueryResult 목록을 시트별로 나눠 하나의 엑셀 파일로 저장한다."""
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()
    for res in results:
        ws = wb.create_sheet(_unique_sheet_name(res.tab, used_names))
        if res.error is not None:
            _write_error(ws, res)
        else:
            _write_table(ws, res.columns, res.rows)
    if not wb.sheetnames:
        wb.create_sheet("결과 없음")
    wb.save(path)


def _write_table(ws, columns, rows):
    for c, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c, value=_clean(col_name))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = HEADER_ALIGN

    n_rows = min(len(rows), EXCEL_MAX_ROWS - 1)
    if n_rows < len(rows):
        log.warning(
            "'%s' 시트: 엑셀 행 한도 초과로 %d행 중 %d행만 기록",
            ws.title, len(rows), n_rows,
        )
    for r, row in enumerate(rows[:n_rows], 2):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=_clean(value))
            cell.border = BORDER
            cell.alignment = CELL_ALIGN

    if columns:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{n_rows + 1}"
        ws.freeze_panes = "A2"
        _autosize_columns(ws, columns, rows[:WIDTH_SAMPLE_ROWS])


def _write_error(ws, res):
    labels = ("오류", "SQL")
    values = (res.error, res.sql)
    for r, (label, value) in enumerate(zip(labels, values), 1):
        head = ws.cell(row=r, column=1, value=label)
        head.font = ERROR_FONT if r == 1 else HEADER_FONT
        head.border = BORDER
        head.alignment = HEADER_ALIGN
        body = ws.cell(row=r, column=2, value=_clean(value))
        body.border = BORDER
        body.alignment = CELL_ALIGN
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 100


def _clean(value):
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _visual_width(value):
    """전각(한글·한자·기호 W/F) 문자를 2칸으로 계산한 표시 폭. 여러 줄이면 가장 긴 줄 기준."""
    text = value if isinstance(value, str) else str(value)
    lines = text.splitlines() or [""]
    return max(
        sum(2 if unicodedata.east_asian_width(ch) in "WF" else 1 for ch in line)
        for line in lines
    )


def _autosize_columns(ws, columns, sample_rows):
    for c, col_name in enumerate(columns, 1):
        width = _visual_width(col_name)
        for row in sample_rows:
            if c <= len(row) and row[c - 1] is not None:
                width = max(width, _visual_width(row[c - 1]))
        ws.column_dimensions[get_column_letter(c)].width = max(
            MIN_COL_WIDTH, min(width + 2, MAX_COL_WIDTH)
        )


def _unique_sheet_name(name, used):
    base = _INVALID_SHEET_CHARS.sub("_", str(name)).strip().strip("'") or "Sheet"
    base = base[:MAX_SHEET_NAME]
    candidate, n = base, 1
    while candidate.lower() in used:
        n += 1
        suffix = f"_{n}"
        candidate = base[: MAX_SHEET_NAME - len(suffix)] + suffix
    used.add(candidate.lower())
    return candidate

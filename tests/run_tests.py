"""파서·엑셀 서식·유니코드 보존·검증기 회귀 테스트 (DB 불필요).

실행: python tests/run_tests.py
"""

import datetime
import decimal
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from db import QueryResult  # noqa: E402
from excel_writer import write_workbook  # noqa: E402
from sql_script import parse_script  # noqa: E402
from verifier import verify_workbook  # noqa: E402

from openpyxl import load_workbook  # noqa: E402

failures = []


def check(cond, msg):
    print(("  ok  " if cond else "FAIL  ") + msg)
    if not cond:
        failures.append(msg)


def test_parser():
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "s1.sql"
        p.write_text(
            "-- @tab: 사용자 목록\nSELECT * FROM users;\n\n"
            "-- @tab: 주문\nSELECT 1;\n\nSELECT 2;\n",
            encoding="utf-8",
        )
        qs = parse_script(p)
        check(len(qs) == 3, f"파서: 기본 구분자로 쿼리 3개 (실제 {len(qs)})")
        check(qs[0].tab == "사용자 목록", f"파서: 한글 탭 이름 (실제 {qs[0].tab!r})")
        check(qs[2].tab == "Query3", f"파서: 탭 자동 번호 (실제 {qs[2].tab!r})")

        p2 = Path(td) / "s2.sql"
        p2.write_text(
            "-- @separator: @@\n"
            "-- @tab: 블록\nBEGIN\n  x := 1;\nEND;\n@@\n"
            "-- @tab: 일반\nSELECT 1 FROM dual\n@@\n",
            encoding="utf-8-sig",
        )
        qs2 = parse_script(p2)
        check(len(qs2) == 2, f"파서: 커스텀 구분자 @@ (실제 {len(qs2)})")
        check("x := 1;" in qs2[0].sql, "파서: 본문의 ; 보존")
        check(qs2[0].tab == "블록", "파서: BOM 파일 처리")


def _sample_results():
    r1 = QueryResult("한글 탭 ★", "SELECT ...")
    r1.columns = ["이름", "기호", "금액", "가입일"]
    r1.rows = [
        ["홍길동", "★→♥♦ 😀 ∑∫≠ ₩€¥", decimal.Decimal("12345.67"), datetime.date(2024, 1, 15)],
        ["李四\n둘째 줄", "ügöß Ελληνικά Русский", None, None],
        ["=SUM(A1:A2)", "수식 주입 방지", 0, datetime.datetime(2024, 3, 1, 9, 30)],
    ]
    r2 = QueryResult("오류탭[금지/문자]", "SELECT * FROM 없는테이블")
    r2.error = "ORA-00942: 테이블 또는 뷰가 존재하지 않습니다"
    return [r1, r2]


def test_excel_and_verifier():
    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "out.xlsx"
        results = _sample_results()
        sheet_map = write_workbook(out, results)

        wb = load_workbook(out)
        ws = wb[sheet_map[0][0]]
        check(ws["A1"].font.bold and ws["A1"].alignment.wrap_text, "서식: 헤더 볼드+줄바꿈")
        check(ws.auto_filter.ref == "A1:D4", f"서식: 자동필터 (실제 {ws.auto_filter.ref})")
        check(ws.freeze_panes == "A2", "서식: 첫 행 고정")
        check(ws["B2"].border.left.style == "thin", "서식: 검정 실선 테두리")
        check(ws["B2"].value == "★→♥♦ 😀 ∑∫≠ ₩€¥", "유니코드: 기호·이모지 보존")
        check(ws["A3"].value == "李四\n둘째 줄", "유니코드: 한자·개행 보존")
        check(ws["A4"].data_type == "s" and ws["A4"].value == "=SUM(A1:A2)",
              "안전: '=' 값이 수식이 아닌 텍스트로 저장")
        wb.close()

        mismatches = verify_workbook(out, sheet_map)
        check(mismatches == [], f"검증기: 정상 파일 통과 (불일치 {mismatches})")

        # 파일을 고의로 훼손하면 검증기가 잡아내야 한다
        wb = load_workbook(out)
        ws = wb[sheet_map[0][0]]
        ws["B2"] = "★→♥♦ ? ∑∫≠ ₩€¥"        # 이모지가 ?로 깨진 상황 재현
        ws["C4"] = 999                        # 숫자 변조
        ws.delete_rows(3)                     # 행 누락
        wb.save(out)
        wb.close()
        mismatches = verify_workbook(out, sheet_map)
        check(len(mismatches) >= 3, f"검증기: 훼손 감지 (감지 {len(mismatches)}건)")
        check(any("2행 2열" in m for m in mismatches), "검증기: 깨진 유니코드 위치 특정")
        check(any("누락" in m or "행" in m for m in mismatches), "검증기: 행 누락 감지")


def test_verifier_type_rules():
    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "t.xlsx"
        r = QueryResult("타입", "SELECT ...")
        r.columns = ["빈문자열", "제어문자", "불리언", "큰수"]
        r.rows = [["", "a\x02b", True, decimal.Decimal("9999999999.123456")]]
        sheet_map = write_workbook(out, [r])
        mismatches = verify_workbook(out, sheet_map)
        check(mismatches == [],
              f"검증기: 정당한 변환(빈문자열→빈 셀, 제어문자 제거 등) 허용 (불일치 {mismatches})")


test_parser()
test_excel_and_verifier()
test_verifier_type_rules()

print()
print("결과:", "전체 통과" if not failures else f"{len(failures)}개 실패")
sys.exit(1 if failures else 0)
